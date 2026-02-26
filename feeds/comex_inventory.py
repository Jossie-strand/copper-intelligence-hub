"""
feeds/comex_inventory.py

Fetches COMEX Copper Vault Inventory XLS from CME Group.
Parses per-warehouse registered/eligible/total (prev + today),
writes full detail to COMEX tab and updates shared Dashboard.

XLS structure (cols 0-8):
  Col 0: Label (warehouse name, Registered, Eligible, Total, grand totals)
  Col 2: Prev Total (st)
  Col 3: Received (st)
  Col 4: Withdrawn (st)
  Col 5: Net Change (st)
  Col 6: Adjustment (st)
  Col 7: Total Today (st)
"""

import os
import json
import datetime
import requests
import openpyxl
import gspread
from google.oauth2.service_account import Credentials
from dashboard import write_exchange, ensure_headers

# ── CONFIG ────────────────────────────────────────────────────
CME_URL       = "https://www.cmegroup.com/delivery_reports/Copper_Stocks.xls"
SHEET_NAME    = "3-exchange-inventory-tracker"
TAB_COMEX     = "COMEX"
TAB_DASHBOARD = "Dashboard"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]

ST_TO_MT = 0.907185

WAREHOUSES = [
    "BALTIMORE", "DETROIT", "EL PASO",
    "NEW ORLEANS", "OWENSBORO", "SALT LAKE CITY", "TUCSON"
]

# COMEX tab headers — totals + per-warehouse prev/today for registered, eligible, total
COMEX_HEADERS = [
    "Date", "Report Date", "Activity Date",
    "Registered (st)", "Eligible (st)", "Total (st)", "Total (mt)",
    "BALTIMORE Prev Total", "BALTIMORE Total Today",
    "DETROIT Prev Total", "DETROIT Total Today",
    "EL PASO Prev Total", "EL PASO Total Today",
    "NEW ORLEANS Prev Total", "NEW ORLEANS Total Today",
    "OWENSBORO Prev Total", "OWENSBORO Total Today",
    "SALT LAKE CITY Prev Total", "SALT LAKE CITY Total Today",
    "TUCSON Prev Total", "TUCSON Total Today",
    "Source"
]


# ── FETCH ─────────────────────────────────────────────────────
def fetch_xls():
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0",
        "Accept": "application/vnd.ms-excel,*/*",
        "Referer": "https://www.cmegroup.com/"
    }
    resp = requests.get(CME_URL, headers=headers, timeout=30)
    resp.raise_for_status()
    print(f"Fetched XLS: HTTP {resp.status_code}, {len(resp.content)} bytes")
    return resp.content


# ── PARSE ─────────────────────────────────────────────────────
def _cell(ws, row_idx, col_idx):
    """Safe cell reader — returns value or None."""
    try:
        val = ws.cell(row=row_idx + 1, column=col_idx + 1).value
        return val
    except Exception:
        return None


def _num(val):
    """Convert cell to float, None if blank/text."""
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def parse_xls(content):
    import io
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    result = {
        "report_date":   "",
        "activity_date": "",
        "registered_st": None,
        "eligible_st":   None,
        "total_st":      None,
        "total_mt":      None,
        "registered_mt": None,
        "eligible_mt":   None,
        "warehouses": {
            wh: {"reg_prev": None, "reg_today": None,
                 "elig_prev": None, "elig_today": None,
                 "total_prev": None, "total_today": None}
            for wh in WAREHOUSES
        }
    }

    rows = list(ws.iter_rows(values_only=True))

    # Log first 20 rows for debugging
    for i, row in enumerate(rows[:20]):
        print(f"Row {i}: {list(row)}")

    current_warehouse = None

    for i, row in enumerate(rows):
        label = str(row[0]).strip().upper() if row[0] else ""

        # Dates — row 7: Report Date, row 8: Activity Date
        if "REPORT DATE" in label and row[6]:
            val = str(row[6]).replace("Report Date:", "").strip()
            result["report_date"] = val
        if "ACTIVITY DATE" in label and row[6]:
            val = str(row[6]).replace("Activity Date:", "").strip()
            result["activity_date"] = val
        # Also check col 6 directly for date strings
        if row[6] and "Report Date:" in str(row[6]):
            result["report_date"] = str(row[6]).replace("Report Date:", "").strip()
        if row[6] and "Activity Date:" in str(row[6]):
            result["activity_date"] = str(row[6]).replace("Activity Date:", "").strip()

        # Warehouse header row
        if label in WAREHOUSES:
            current_warehouse = label
            continue

        # Per-warehouse data rows
        if current_warehouse and label in ("REGISTERED (WARRANTED)", "REGISTERED"):
            result["warehouses"][current_warehouse]["reg_prev"]  = _num(row[2])
            result["warehouses"][current_warehouse]["reg_today"] = _num(row[7])

        if current_warehouse and label in ("ELIGIBLE (NON-WARRANTED)", "ELIGIBLE"):
            result["warehouses"][current_warehouse]["elig_prev"]  = _num(row[2])
            result["warehouses"][current_warehouse]["elig_today"] = _num(row[7])

        if current_warehouse and label == "TOTAL" and _num(row[2]) is not None:
            result["warehouses"][current_warehouse]["total_prev"]  = _num(row[2])
            result["warehouses"][current_warehouse]["total_today"] = _num(row[7])

        # Grand totals
        if "TOTAL REGISTERED" in label:
            result["registered_st"] = _num(row[7])
            print(f"  → registered_st = {result['registered_st']}")

        if "TOTAL ELIGIBLE" in label:
            result["eligible_st"] = _num(row[7])
            print(f"  → eligible_st = {result['eligible_st']}")

        if label == "TOTAL COPPER":
            result["total_st"] = _num(row[7])
            print(f"  → total_st = {result['total_st']}")

    # Fallback
    if result["total_st"] is None and result["registered_st"] and result["eligible_st"]:
        result["total_st"] = result["registered_st"] + result["eligible_st"]

    if result["total_st"]:
        result["total_mt"] = round(result["total_st"] * ST_TO_MT)
    if result["registered_st"]:
        result["registered_mt"] = round(result["registered_st"] * ST_TO_MT)
    if result["eligible_st"]:
        result["eligible_mt"] = round(result["eligible_st"] * ST_TO_MT)

    return result


# ── GOOGLE SHEETS ─────────────────────────────────────────────
def get_sheet_client():
    creds_dict = json.loads(os.environ["GOOGLE_SERVICE_ACCOUNT_JSON"])
    creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    return gspread.authorize(creds)


def ensure_comex_headers(tab):
    first = tab.row_values(1)
    if not first or first[0] != "Date":
        tab.insert_row(COMEX_HEADERS, 1)
        print("✅ COMEX headers written")


def calc_comex_change(tab, total_mt):
    """Delta from most recent prior row's Total (mt) — col G (col 7)."""
    if total_mt is None:
        return None
    all_vals = tab.col_values(7)
    for v in reversed(all_vals[1:]):
        try:
            prev = float(str(v).replace(",", ""))
            return round(total_mt - prev, 0)
        except (ValueError, TypeError):
            continue
    return None


def write_to_sheet(data):
    client    = get_sheet_client()
    book      = client.open(SHEET_NAME)
    comex_tab = book.worksheet(TAB_COMEX)
    dash_tab  = book.worksheet(TAB_DASHBOARD)

    today = datetime.date.today().isoformat()

    ensure_comex_headers(comex_tab)

    # Duplicate check on activity_date (col C = col 3)
    existing = comex_tab.col_values(3)
    already_logged = data["activity_date"] and data["activity_date"] in existing

    change_mt = calc_comex_change(comex_tab, data["total_mt"])

    if already_logged:
        print(f"Duplicate: activity_date {data['activity_date']} already in COMEX tab. Skipping tab write.")
    else:
        # Build per-warehouse cells: Prev Total, Total Today only
        wh_cells = []
        for wh in WAREHOUSES:
            w = data["warehouses"][wh]
            wh_cells += [
                w["total_prev"]  or "",
                w["total_today"] or "",
            ]

        comex_row = [
            today,
            data["report_date"],
            data["activity_date"],
            data["registered_st"] or "",
            data["eligible_st"]   or "",
            data["total_st"]      or "",
            data["total_mt"]      or "",
            *wh_cells,
            CME_URL
        ]
        comex_tab.append_row(comex_row, value_input_option="USER_ENTERED")
        print(f"✅ COMEX tab: {data['total_mt']} mt | registered {data['registered_mt']} mt | change {change_mt} mt")
        for wh in WAREHOUSES:
            w = data["warehouses"][wh]
            print(f"   {wh}: reg={w['reg_today']} elig={w['elig_today']} total={w['total_today']}")

    # Dashboard always runs — keyed on activity_date
    data_date = today
    if data["activity_date"]:
        try:
            data_date = datetime.datetime.strptime(
                data["activity_date"], "%m/%d/%Y"
            ).strftime("%Y-%m-%d")
        except ValueError:
            pass

    ensure_headers(dash_tab)
    write_exchange(
        dash_tab, data_date, "COMEX",
        total_mt=data["total_mt"],
        change_mt=change_mt,
        extras={
            "registered": data["registered_mt"],
            "eligible":   data["eligible_mt"],
        }
    )


# ── MAIN ──────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print(f"COMEX Copper Feed — {datetime.datetime.utcnow().isoformat()}Z")
    print("=" * 60)

    try:
        content = fetch_xls()
        data    = parse_xls(content)
        print(f"\nParsed: total_st={data['total_st']}, total_mt={data['total_mt']}, "
              f"registered_mt={data['registered_mt']}, eligible_mt={data['eligible_mt']}")

        if not data["total_st"]:
            raise ValueError("Parse failed — total_st is None")

        write_to_sheet(data)
        print("\n✅ Done.")

    except Exception as e:
        print(f"\n❌ Error: {e}")
        raise


if __name__ == "__main__":
    main()
