"""
Copper Intelligence Hub — Automated Inventory Data Pull
Phase 02: Optional automation script

This script attempts to fetch daily/weekly inventory data from:
  - LME (lme.com) — daily copper warrants
  - CME Group (cmegroup.com) — COMEX copper stocks
  - SHFE (shfe.com.cn) — weekly copper stocks

Then auto-populates the 3-exchange-inventory-tracker.xlsx

REQUIREMENTS:
  pip install requests pandas openpyxl beautifulsoup4 lxml selenium

NOTE: These sites may require browser automation (Selenium) or API keys
      for reliable data access. This script provides the framework;
      you may need to adjust selectors if site layouts change.

SCHEDULE: Run daily (LME/COMEX) or weekly on Fridays (SHFE)
  - Mac/Linux crontab:  0 8 * * 1-5 python3 /path/to/04-inventory-data-pull.py
  - Windows Task Scheduler: create task pointing to this script
"""

import requests
import pandas as pd
from datetime import datetime, date
from pathlib import Path
import json
import time
import sys

# ── CONFIG ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
TRACKER_PATH = SCRIPT_DIR / "03-3exchange-inventory-tracker.xlsx"
LOG_PATH = SCRIPT_DIR / "_pull_log.json"

TODAY = date.today()
TODAY_STR = TODAY.strftime("%Y-%m-%d")

HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/120.0.0.0 Safari/537.36")
}


# ── LOGGING ────────────────────────────────────────────────────────────────────
def log(msg: str, level: str = "INFO"):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] [{level}] {msg}")


def save_log(data: dict):
    history = []
    if LOG_PATH.exists():
        with open(LOG_PATH) as f:
            history = json.load(f)
    history.append({"timestamp": datetime.now().isoformat(), **data})
    with open(LOG_PATH, "w") as f:
        json.dump(history[-90:], f, indent=2)  # keep 90 days


# ── LME DATA ────────────────────────────────────────────────────────────────────
def fetch_lme() -> dict | None:
    """
    LME publishes daily stock reports as PDF/Excel via:
    https://www.lme.com/en/Market-Data/Reports-and-data/Monthly-reports

    For automated access, the LME Data API (paid) is recommended.
    This function demonstrates a free fallback using the public warehouse report.

    Returns: dict with keys: warrants, cancelled_warrants, date
    """
    log("Fetching LME inventory data...")

    # Attempt: LME public warehouse stocks endpoint
    # NOTE: The actual endpoint may require LME login or change over time.
    # Adjust URL as needed after checking lme.com current structure.
    url = "https://www.lme.com/api/v1/market-data/warehouse-stocks/copper"

    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code == 200:
            data = r.json()
            # Parse based on actual response structure
            warrants = data.get("totalWarrants") or data.get("warrants")
            cancelled = data.get("cancelledWarrants") or data.get("cancelled")
            if warrants:
                log(f"LME: Warrants={warrants:,} t, Cancelled={cancelled:,} t")
                return {"warrants": warrants, "cancelled_warrants": cancelled,
                        "date": TODAY_STR, "source": "LME API"}
        log(f"LME API returned {r.status_code} — falling back to manual entry", "WARN")
    except Exception as e:
        log(f"LME fetch failed: {e}", "WARN")

    # FALLBACK: prompt user for manual entry
    log(">>> MANUAL ENTRY NEEDED — go to lme.com and enter values below", "MANUAL")
    return None


# ── COMEX DATA ──────────────────────────────────────────────────────────────────
def fetch_comex() -> dict | None:
    """
    CME Group publishes COMEX copper warehouse stocks daily.
    URL: https://www.cmegroup.com/delivery_reports/MetalsIssuesAndStopsReport.pdf
    Or via CME DataMine API (paid).

    Returns: dict with eligible, registered stocks
    """
    log("Fetching COMEX inventory data...")

    # CME has a public metals delivery report
    # Try parsing the daily report (HTML version)
    url = "https://www.cmegroup.com/CmeWS/mvc/Warehouse/Report?instrumentId=39"

    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code == 200:
            # Attempt to parse JSON response
            data = r.json()
            eligible = None
            registered = None
            # Navigate actual structure — adjust keys to match real response
            for item in data.get("data", []):
                if "COPPER" in str(item).upper():
                    eligible = item.get("eligible", item.get("eligibleTotal"))
                    registered = item.get("registered", item.get("registeredTotal"))
                    break
            if eligible is not None:
                log(f"COMEX: Eligible={eligible:,}, Registered={registered:,}")
                return {"eligible": eligible, "registered": registered,
                        "date": TODAY_STR, "source": "CME API"}
    except Exception as e:
        log(f"COMEX fetch failed: {e}", "WARN")

    log(">>> MANUAL ENTRY NEEDED — go to cmegroup.com → Metals → Delivery → Copper", "MANUAL")
    return None


# ── SHFE DATA ───────────────────────────────────────────────────────────────────
def fetch_shfe() -> dict | None:
    """
    SHFE publishes weekly copper inventory on Fridays.
    URL: http://www.shfe.com.cn/en/statements/MarketData/

    Returns: dict with weekly stocks
    """
    if TODAY.weekday() != 4:  # Only run on Fridays
        log("SHFE: Not Friday — skipping weekly pull", "SKIP")
        return None

    log("Fetching SHFE weekly inventory (Friday run)...")

    try:
        url = "http://www.shfe.com.cn/data/dailydata/kx/kx20" + TODAY.strftime("%m%d") + ".dat"
        r = requests.get(url, headers=HEADERS, timeout=20)
        if r.status_code == 200:
            data = r.json()
            for item in data.get("o_cursor", []):
                if item.get("VARNAME") == "cu":
                    stocks = item.get("WH")
                    log(f"SHFE: Copper stocks = {stocks:,} t")
                    return {"stocks": stocks, "date": TODAY_STR, "source": "SHFE"}
    except Exception as e:
        log(f"SHFE fetch failed: {e}", "WARN")

    log(">>> MANUAL ENTRY NEEDED — go to shfe.com.cn → Market Data → Inventory", "MANUAL")
    return None


# ── WRITE TO EXCEL ─────────────────────────────────────────────────────────────
def update_tracker(lme: dict, comex: dict, shfe: dict):
    """Append fetched data to the tracking spreadsheet."""
    if not TRACKER_PATH.exists():
        log(f"Tracker not found at {TRACKER_PATH}", "ERROR")
        return

    from openpyxl import load_workbook
    wb = load_workbook(TRACKER_PATH)

    # LME sheet
    if lme:
        ws = wb["LME"]
        next_row = ws.max_row + 1
        ws.cell(next_row, 1, TODAY_STR)
        ws.cell(next_row, 2, lme["warrants"])
        ws.cell(next_row, 3, lme["cancelled_warrants"])
        log(f"LME data written to row {next_row}")

    # COMEX sheet
    if comex:
        ws = wb["COMEX"]
        next_row = ws.max_row + 1
        ws.cell(next_row, 1, TODAY_STR)
        ws.cell(next_row, 2, comex["eligible"])
        ws.cell(next_row, 3, comex["registered"])
        log(f"COMEX data written to row {next_row}")

    # SHFE sheet
    if shfe:
        ws = wb["SHFE"]
        next_row = ws.max_row + 1
        ws.cell(next_row, 1, TODAY_STR)
        ws.cell(next_row, 2, shfe["stocks"])
        log(f"SHFE data written to row {next_row}")

    wb.save(TRACKER_PATH)
    log(f"Tracker saved: {TRACKER_PATH}")


# ── MAIN ────────────────────────────────────────────────────────────────────────
def main():
    log("=" * 60)
    log(f"Copper Hub — Inventory Pull — {TODAY_STR}")
    log("=" * 60)

    results = {}

    lme = fetch_lme()
    results["lme"] = "ok" if lme else "manual_required"

    comex = fetch_comex()
    results["comex"] = "ok" if comex else "manual_required"

    shfe = fetch_shfe()
    results["shfe"] = "ok" if shfe else ("skipped" if TODAY.weekday() != 4 else "manual_required")

    if any(v == "ok" for v in results.values()):
        update_tracker(lme, comex, shfe)

    save_log({"date": TODAY_STR, "results": results})

    log("\n── Summary ──────────────────────────────────────")
    for source, status in results.items():
        icon = "✓" if status == "ok" else ("—" if status == "skipped" else "⚠ MANUAL")
        log(f"  {source.upper():8} {icon}")

    if any(v == "manual_required" for v in results.values()):
        log("\n⚠  Some sources need manual entry. Open the tracker and fill highlighted rows.")
        log("   LME:   lme.com → Market Data → Reports & Data → Monthly Reports")
        log("   COMEX: cmegroup.com → Copper → Delivery → Warehouse Stocks")
        log("   SHFE:  shfe.com.cn → Market Data → Inventory")


if __name__ == "__main__":
    main()
